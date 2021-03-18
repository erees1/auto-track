from argparse import ArgumentParser, FileType
from openpyxl import load_workbook
import datetime
import sys
import re
from src.utils import load_config
from dataclasses import asdict
import csv


def get_data_from_sheet(
    ws, desc_col="F", cat_col="E", val_col="D", date_col="C", start_row=69, **kwargs
):
    output_data = []
    row = start_row
    while True:
        val = ws[val_col + str(row)].value
        if val is None:
            break
        desc = clean_data(ws[desc_col + str(row)].value)
        date = ws[date_col + str(row)].value
        cat = ws[cat_col + str(row)].value
        output_data.append({"date": date, "val": val, "desc": desc, "cat": cat})
        row += 1
    return output_data


def get_data_from_workbook(
    infile, desc_col="F", cat_col="E", val_col="D", date_col="C", start_row=69, **kwargs
):
    wb = load_workbook(infile)
    sheets = [datetime.date(2008, i + 1, 1).strftime("%B") for i in range(12)]
    output_data = []
    for sheet in sheets:
        if sheet in wb:
            output_data.extend(
                get_data_from_sheet(wb[sheet], desc_col, cat_col, val_col, date_col, start_row)
            )
    return output_data


def clean_data(t):
    if t is None:
        return ""
    elif t is not str:
        t = str(t)
    t = re.sub(" +", " ", t)
    t = re.sub(",", "", t)
    t = t.lower()
    return t


def main():
    parser = ArgumentParser()
    parser.add_argument("infile", nargs="?", type=str)
    parser.add_argument("outfile", nargs="?", type=FileType("w"), default=sys.stdout)
    parser.add_argument("--config_path", default="DEFAULT", help="path to config file")
    args = parser.parse_args()

    wb_config = load_config(args.config_path).wb
    output = get_data_from_workbook(args.infile, **asdict(wb_config))
    writer = csv.DictWriter(args.outfile, fieldnames=output[0].keys())
    writer.writeheader()
    for data in output:
        writer.writerow(data)


if __name__ == "__main__":
    main()
