import argparse
from openpyxl import load_workbook
from datetime import datetime
from utils import to_num

# from extract_features import get_data_from_sheet
from utils import GlobalConfig, load_config, get_providers


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--csv_path", help="Path to csv file to import")
    parser.add_argument("--tracker_path", help="Path to destination track to copy imports into")
    parser.add_argument("--provider", help="name of config type, e.g. monzo")
    parser.add_argument("--config", help="Path to config file if not in default config folder")
    args = parser.parse_args()

    if args.csv_path is None:
        args.csv_path = input("Select csv file to import:\n")
    if args.tracker_path is None:
        args.tracker_path = input("Select path to tracker excel file:\n")
    if args.provider is None:
        available_provders = get_providers(args.config)
        args.provider = input(
            f"Please enter the name of the config type, available_provders are {available_provders}:\n"
        )

    import_transactions(args.csv_path, args.tracker_path, load_config(args.config, args.provider))


def get_ammount(lsplit, csv_config):
    ammount = lsplit[csv_config.val_field]
    am = float(ammount) if csv_config.ammount_is_negative else -float(ammount)
    return am


def get_desc(lsplit, csv_config):

    if isinstance(csv_config.desc_field, list):
        desc = " - ".join([lsplit[i] for i in csv_config.desc_field])
    else:
        desc = lsplit[csv_config.desc_field]
    return desc


def get_date(lsplit, csv_config):
    date = datetime.strptime(lsplit[csv_config.date_field], csv_config.date_frmt)
    return date


def has_skip(desc, csv_config):

    # Skip entry if skip words are present
    for skip_word in csv_config.skip_words:
        if skip_word.lower() in desc.lower():
            skip = True
            # print(f"found {skip_word} in {desc}")
            break
        else:
            skip = False
    return skip


def import_transactions(fp, tp, config: GlobalConfig):
    keep_vba = True if "xlsm" in tp else False

    wb = load_workbook(tp, keep_vba=keep_vba)
    # exisiting_data = None

    wb_config = config.wb
    csv_config = config.csv

    with open(fp) as f:
        lines = f.readlines()
        if csv_config.has_header_row:
            s = 1
        else:
            s = 0
        for line in lines[s:]:
            lsplit = line.split(",")
            # Get the fields
            ammount = get_ammount(lsplit, csv_config)
            desc = get_desc(lsplit, csv_config)
            date = get_date(lsplit, csv_config)

            if has_skip(desc, csv_config):
                continue

            # Copy procedure
            ws = wb[wb_config.data_sheet]
            row = get_first_empty_row(ws, wb_config.val_col)
            ws[wb_config.val_col + str(row)].value = ammount
            ws[wb_config.desc_col + str(row)].value = desc
            ws[wb_config.date_col + str(row)].value = date
            ws[wb_config.source_col + str(row)].value = csv_config.source.capitalize()
            row += 1

    wb.save(tp)
    print("Updated")


def get_first_empty_row(ws, col, start_row=1):
    x = start_row
    col_ix = to_num(col) + 1
    while True:
        if (
            ws.cell(
                row=x,
                column=col_ix,
            ).value
            is None
        ):
            return x
        else:
            x += 1


if __name__ == "__main__":
    main()
